#!/usr/bin/env python3
"""openpyxl runner for competitive real-workbook benchmarks."""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import re
import time
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

import openpyxl
from memory_metrics import memory_baseline, sample_with_memory
from package_fingerprint import (
    parse_relationships,
    read_text,
    resolve_path,
    roundtrip_feature_assertions,
    roundtrip_package_assertions,
)


def hash_lines(lines: list[str]) -> str:
    digest = hashlib.sha256()
    for line in sorted(lines):
        digest.update(f"{len(line)}:".encode("utf-8"))
        digest.update(line.encode("utf-8"))
        digest.update(b"\n")
    return digest.hexdigest()


def cell_ref(sheet_name: str, cell: Any) -> str:
    return f"{sheet_name}!{openpyxl.utils.get_column_letter(cell.column)}{cell.row}"


def cell_ref_from_coords(sheet_name: str, row: int, column: int) -> str:
    return f"{sheet_name}!{openpyxl.utils.get_column_letter(column)}{row}"


def formula_text(value: Any) -> str | None:
    if not isinstance(value, str) or not value.startswith("="):
        return None
    return value[1:]


def canonical_number(value: Any) -> str:
    number = float(value)
    if number == 0:
        return "0"
    if number.is_integer():
        return str(int(number))
    return format(number, ".15g")


def scalar_payload(cell: Any) -> str:
    if cell.data_type == "f":
        return "empty"
    if cell.value is None:
        return "empty"
    if cell.data_type == "b":
        return f"b:{'true' if bool(cell.value) else 'false'}"
    if cell.data_type == "e":
        return f"e:{cell.value}"
    if getattr(cell, "is_date", False):
        return f"s:{cell.value.isoformat() if hasattr(cell.value, 'isoformat') else cell.value}"
    if isinstance(cell.value, (int, float)):
        return f"n:{canonical_number(cell.value)}"
    return f"s:{cell.value}"


def used_range(sheet_name: str, cells: list[Any]) -> str:
    if not cells:
        return f"{sheet_name}!empty"
    min_row = min(cell.row for cell in cells)
    min_col = min(cell.column for cell in cells)
    max_row = max(cell.row for cell in cells)
    max_col = max(cell.column for cell in cells)
    return (
        f"{sheet_name}!"
        f"{openpyxl.utils.get_column_letter(min_col)}{min_row}:"
        f"{openpyxl.utils.get_column_letter(max_col)}{max_row}"
    )


def used_range_from_coords(sheet_name: str, coords: list[tuple[int, int]]) -> str:
    if not coords:
        return f"{sheet_name}!empty"
    min_row = min(row for row, _column in coords)
    min_col = min(column for _row, column in coords)
    max_row = max(row for row, _column in coords)
    max_col = max(column for _row, column in coords)
    return (
        f"{sheet_name}!"
        f"{openpyxl.utils.get_column_letter(min_col)}{min_row}:"
        f"{openpyxl.utils.get_column_letter(max_col)}{max_row}"
    )


def cell_coords(cell: Any) -> tuple[int, int] | None:
    row = getattr(cell, "row", None)
    column = getattr(cell, "column", None)
    if isinstance(row, int) and isinstance(column, int):
        return (row, column)
    return None


def shape_assertions_from_workbook(
    workbook: Any, prefix: str = ""
) -> dict[str, str | int | bool | None]:
    if getattr(workbook, "read_only", False):
        return streaming_shape_assertions_from_workbook(workbook, prefix)

    def key(name: str) -> str:
        return f"{prefix}{name[0].upper()}{name[1:]}" if prefix else name

    cell_count = 0
    physical_cell_count = 0
    formula_count = 0
    sheet_names: list[str] = [sheet.title for sheet in workbook.worksheets]
    used_ranges: list[str] = []
    physical_used_ranges: list[str] = []
    semantic_cell_refs: list[str] = []
    semantic_cell_values: list[str] = []
    formula_texts: list[str] = []
    for sheet in workbook.worksheets:
        physical_cells = list(sheet._cells.values())
        semantic_cells = [
            cell for cell in physical_cells if cell.value is not None or cell.data_type == "f"
        ]
        physical_cell_count += len(physical_cells)
        cell_count += len(semantic_cells)
        for cell in semantic_cells:
            ref = cell_ref(sheet.title, cell)
            semantic_cell_refs.append(ref)
            semantic_cell_values.append(f"{ref}\t{scalar_payload(cell)}")
            formula = formula_text(cell.value)
            if formula is not None:
                formula_count += 1
                formula_texts.append(f"{ref}={formula}")
        used_ranges.append(used_range(sheet.title, semantic_cells))
        physical_used_ranges.append(used_range(sheet.title, physical_cells))
    return {
        key("sheetCount"): len(workbook.worksheets),
        key("sheetNamesHash"): hash_lines([f"{index}:{name}" for index, name in enumerate(sheet_names)]),
        key("cellCount"): cell_count,
        key("physicalCellCount"): physical_cell_count,
        key("formulaCount"): formula_count,
        key("usedRangeCount"): len(used_ranges),
        key("firstUsedRange"): used_ranges[0] if used_ranges else None,
        key("firstPhysicalUsedRange"): physical_used_ranges[0] if physical_used_ranges else None,
        key("usedRangesHash"): hash_lines(used_ranges),
        key("physicalUsedRangesHash"): hash_lines(physical_used_ranges),
        key("semanticCellRefsHash"): hash_lines(semantic_cell_refs),
        key("semanticCellValuesHash"): hash_lines(semantic_cell_values),
        key("formulaTextHash"): hash_lines(formula_texts),
    }


def streaming_shape_assertions_from_workbook(
    workbook: Any, prefix: str = ""
) -> dict[str, str | int | bool | None]:
    return streaming_shape_assertions_from_data(streaming_shape_data_from_workbook(workbook), prefix)


def streaming_shape_data_from_workbook(workbook: Any) -> dict[str, Any]:
    cell_count = 0
    formula_count = 0
    sheet_names: list[str] = [sheet.title for sheet in workbook.worksheets]
    used_ranges: list[str] = []
    semantic_cell_refs: list[str] = []
    semantic_cell_values: list[str] = []
    formula_texts: list[str] = []
    for sheet in workbook.worksheets:
        semantic_coords: list[tuple[int, int]] = []
        for row in sheet.iter_rows():
            for cell in row:
                coords = cell_coords(cell)
                if coords is None:
                    continue
                value = getattr(cell, "value", None)
                data_type = getattr(cell, "data_type", None)
                if value is None and data_type != "f":
                    continue
                cell_count += 1
                semantic_coords.append(coords)
                ref = cell_ref_from_coords(sheet.title, coords[0], coords[1])
                semantic_cell_refs.append(ref)
                semantic_cell_values.append(f"{ref}\t{scalar_payload(cell)}")
                formula = formula_text(value)
                if formula is not None:
                    formula_count += 1
                    formula_texts.append(f"{ref}={formula}")
        used_ranges.append(used_range_from_coords(sheet.title, semantic_coords))
    return {
        "__streamingShapeData": True,
        "sheetCount": len(workbook.worksheets),
        "sheetNames": sheet_names,
        "cellCount": cell_count,
        "formulaCount": formula_count,
        "usedRanges": used_ranges,
        "semanticCellRefs": semantic_cell_refs,
        "semanticCellValues": semantic_cell_values,
        "formulaTexts": formula_texts,
    }


def streaming_shape_assertions_from_data(
    data: dict[str, Any], prefix: str = ""
) -> dict[str, str | int | bool | None]:
    def key(name: str) -> str:
        return f"{prefix}{name[0].upper()}{name[1:]}" if prefix else name

    sheet_names = data["sheetNames"]
    used_ranges = data["usedRanges"]
    return {
        key("sheetCount"): data["sheetCount"],
        key("sheetNamesHash"): hash_lines([f"{index}:{name}" for index, name in enumerate(sheet_names)]),
        key("cellCount"): data["cellCount"],
        key("physicalCellCount"): None,
        key("formulaCount"): data["formulaCount"],
        key("usedRangeCount"): len(used_ranges),
        key("firstUsedRange"): used_ranges[0] if used_ranges else None,
        key("firstPhysicalUsedRange"): None,
        key("usedRangesHash"): hash_lines(used_ranges),
        key("physicalUsedRangesHash"): hash_lines([]),
        key("semanticCellRefsHash"): hash_lines(data["semanticCellRefs"]),
        key("semanticCellValuesHash"): hash_lines(data["semanticCellValues"]),
        key("formulaTextHash"): hash_lines(data["formulaTexts"]),
    }


def load_workbook_for_read(path: Path, read_only: bool = False, data_only: bool = False) -> Any:
    return openpyxl.load_workbook(
        path,
        read_only=read_only,
        data_only=data_only,
        keep_vba=path.suffix == ".xlsm",
    )


def metadata_only_assertions(path: Path) -> dict[str, str | int | bool | None]:
    workbook = load_workbook_for_read(path, read_only=True, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
        return {
            "runnerVersion": openpyxl.__version__,
            "metadataOnlyRead": True,
            "sourceSheetCount": len(sheet_names),
            "loadedSheetCount": len(sheet_names),
            "loadedSheetNames": ",".join(sheet_names),
            "hasAllSheets": True,
            "cellsHydrated": False,
            "runnerReadOnly": True,
            "runnerDataOnly": True,
        }
    finally:
        workbook.close()


def shape_assertions(workbook: Any) -> dict[str, str | int | bool | None]:
    try:
        return {"runnerVersion": openpyxl.__version__, **shape_assertions_from_workbook(workbook)}
    finally:
        workbook.close()


def roundtrip_operation(path: Path) -> bytes:
    workbook = load_workbook_for_read(path)
    try:
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()
    finally:
        workbook.close()


def edit_roundtrip_operation(path: Path, sheet_name: str, ref: str, value: float) -> bytes:
    workbook = load_workbook_for_read(path)
    try:
        workbook[sheet_name][ref].value = value
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()
    finally:
        workbook.close()


def roundtrip_assertions(path: Path, data: bytes) -> dict[str, str | int | bool | None]:
    original = path.read_bytes()
    reopened = openpyxl.load_workbook(
        io.BytesIO(data),
        read_only=False,
        data_only=False,
        keep_vba=path.suffix == ".xlsm",
    )
    try:
        shape = shape_assertions_from_workbook(reopened, "roundtrip")
    finally:
        reopened.close()
    return {
        "runnerVersion": openpyxl.__version__,
        "bytes": len(data),
        "byteIdentical": hashlib.sha256(data).digest() == hashlib.sha256(original).digest(),
        **roundtrip_package_assertions(data),
        **roundtrip_feature_assertions(data),
        **shape,
    }


def read_roundtrip_cell_number(data: bytes, sheet_name: str, ref: str) -> float | int | None:
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        workbook_xml = read_text(archive, "xl/workbook.xml")
        workbook_rels_xml = read_text(archive, "xl/_rels/workbook.xml.rels")
        if workbook_xml is None or workbook_rels_xml is None:
            return None
        workbook = ElementTree.fromstring(workbook_xml)
        sheet_rid = None
        for sheet in workbook.iter():
            if not sheet.tag.endswith("sheet") or sheet.attrib.get("name") != sheet_name:
                continue
            sheet_rid = next(
                (
                    value
                    for key, value in sheet.attrib.items()
                    if key.endswith("}id") or key == "r:id"
                ),
                None,
            )
            break
        if sheet_rid is None:
            return None
        relationship = next(
            (
                rel
                for rel in parse_relationships(workbook_rels_xml)
                if rel["id"] == sheet_rid and rel["type"].endswith("/worksheet")
            ),
            None,
        )
        if relationship is None:
            return None
        sheet_xml = read_text(archive, resolve_path("xl/workbook.xml", relationship["target"]))
        if sheet_xml is None:
            return None
        for match in re.finditer(r"<c\b([^>]*)>", sheet_xml):
            attrs = match.group(1) or ""
            ref_match = re.search(r'(?:^|\s)r="([^"]+)"', attrs)
            if ref_match is None or ref_match.group(1) != ref:
                continue
            body_start = match.end()
            body_end = (
                body_start if attrs.rstrip().endswith("/") else sheet_xml.find("</c>", body_start)
            )
            if body_end < body_start:
                return None
            value_match = re.search(r"<v[^>]*>(.*?)</v>", sheet_xml[body_start:body_end], re.S)
            if value_match is None:
                return None
            try:
                observed = float(value_match.group(1))
            except ValueError:
                return None
            return int(observed) if observed.is_integer() else observed
    return None


def edit_roundtrip_assertions(
    path: Path,
    data: bytes,
    sheet_name: str,
    ref: str,
    expected_value: float,
    old_value: float | None,
) -> dict[str, str | int | bool | None]:
    observed = read_roundtrip_cell_number(data, sheet_name, ref)
    return {
        **roundtrip_assertions(path, data),
        "editSheetName": sheet_name,
        "editRef": ref,
        "editOldValue": old_value,
        "editExpectedValue": expected_value,
        "editObservedValue": observed,
        "editCellValueMatches": observed == expected_value,
    }


def run_operation(
    path: Path,
    operation: str,
    read_only: bool,
    data_only: bool,
    edit_sheet: str | None,
    edit_ref: str | None,
    edit_value: float | None,
) -> Any:
    if operation == "read":
        workbook = load_workbook_for_read(path, read_only=read_only, data_only=data_only)
        if read_only:
            try:
                return streaming_shape_data_from_workbook(workbook)
            finally:
                workbook.close()
        return workbook
    if operation == "edit-roundtrip":
        if edit_sheet is None or edit_ref is None or edit_value is None:
            raise ValueError("edit-roundtrip requires --edit-sheet, --edit-ref, and --edit-value")
        return edit_roundtrip_operation(path, edit_sheet, edit_ref, edit_value)
    return roundtrip_operation(path)


def assertions_for_result(
    path: Path,
    operation: str,
    result: Any,
    read_only: bool,
    data_only: bool,
    edit_sheet: str | None,
    edit_ref: str | None,
    edit_value: float | None,
    edit_old_value: float | None,
) -> dict[str, str | int | bool | None]:
    if operation == "read":
        if isinstance(result, dict) and result.get("__streamingShapeData") is True:
            assertions = streaming_shape_assertions_from_data(result)
        else:
            assertions = result if isinstance(result, dict) else shape_assertions(result)
        return {
            "runnerVersion": openpyxl.__version__,
            **assertions,
            "runnerReadOnly": read_only,
            "runnerDataOnly": data_only,
        }
    if operation == "edit-roundtrip":
        if edit_sheet is None or edit_ref is None or edit_value is None:
            raise ValueError("edit-roundtrip requires --edit-sheet, --edit-ref, and --edit-value")
        return edit_roundtrip_assertions(
            path, result, edit_sheet, edit_ref, edit_value, edit_old_value
        )
    return roundtrip_assertions(path, result)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--operation", choices=["read", "roundtrip", "edit-roundtrip"], required=True)
    parser.add_argument("--file", required=True)
    parser.add_argument("--repeat", type=int, default=1)
    parser.add_argument("--warmup", type=int, default=0)
    parser.add_argument("--read-only", action="store_true")
    parser.add_argument("--data-only", action="store_true")
    parser.add_argument("--metadata-only", action="store_true")
    parser.add_argument("--edit-sheet")
    parser.add_argument("--edit-ref")
    parser.add_argument("--edit-value", type=float)
    parser.add_argument("--edit-old-value", type=float)
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()
    if args.operation != "read" and (args.read_only or args.data_only):
        parser.error("--read-only and --data-only are only supported for read operations")
    if args.metadata_only and args.operation != "read":
        parser.error("--metadata-only is only supported for read operations")
    if args.operation == "edit-roundtrip" and (
        args.edit_sheet is None or args.edit_ref is None or args.edit_value is None
    ):
        parser.error("edit-roundtrip requires --edit-sheet, --edit-ref, and --edit-value")

    path = Path(args.file)
    for _ in range(max(0, args.warmup)):
        if args.metadata_only:
            metadata_only_assertions(path)
        else:
            result = run_operation(
                path,
                args.operation,
                args.read_only,
                args.data_only,
                args.edit_sheet,
                args.edit_ref,
                args.edit_value,
            )
            assertions_for_result(
                path,
                args.operation,
                result,
                args.read_only,
                args.data_only,
                args.edit_sheet,
                args.edit_ref,
                args.edit_value,
                args.edit_old_value,
            )
    samples: list[dict[str, float]] = []
    assertions: dict[str, str | int | bool | None] | None = None
    for _ in range(max(1, args.repeat)):
        if args.metadata_only:
            before = memory_baseline()
            start = time.perf_counter()
            assertions = metadata_only_assertions(path)
            duration_ms = (time.perf_counter() - start) * 1000
            samples.append(sample_with_memory(duration_ms, before))
            continue
        before = memory_baseline()
        start = time.perf_counter()
        result = run_operation(
            path,
            args.operation,
            args.read_only,
            args.data_only,
            args.edit_sheet,
            args.edit_ref,
            args.edit_value,
        )
        duration_ms = (time.perf_counter() - start) * 1000
        assertions = assertions_for_result(
            path,
            args.operation,
            result,
            args.read_only,
            args.data_only,
            args.edit_sheet,
            args.edit_ref,
            args.edit_value,
            args.edit_old_value,
        )
        samples.append(sample_with_memory(duration_ms, before))
    payload = {"assertions": assertions or {}, "samples": samples}
    if args.json:
        print(json.dumps(payload, separators=(",", ":")))
    else:
        print(json.dumps(payload, indent=2))


if __name__ == "__main__":
    main()
