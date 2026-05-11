#!/usr/bin/env python3
"""Python high-performance XLSX runner for public matrix benchmark profiles."""

from __future__ import annotations

import argparse
import hashlib
import importlib.metadata as metadata
import io
import json
import re
import tempfile
import time
import zipfile
from pathlib import Path
from typing import Any

import openpyxl
from generated_xlsx_validation import (
    can_validate_generated_workload,
    expected_ordered_values_hash,
    generated_cell_count,
    generated_write_assertions,
)
from memory_metrics import memory_baseline, sample_with_memory

WorkloadName = str
WORKLOAD_CHOICES = [
    "dense-values",
    "mixed-10pct-text",
    "mixed-50pct-text",
    "mixed-closedxml-10text-5number",
    "plain-text",
    "string-heavy",
    "sparse-wide",
    "styles-heavy",
    "formula-heavy",
    "table-heavy",
    "feature-rich",
]


def column_name(index: int) -> str:
    name = ""
    current = index
    while current > 0:
        current, rem = divmod(current - 1, 26)
        name = chr(65 + rem) + name
    return name or "A"


def hash_lines(lines: list[str]) -> str:
    digest = hashlib.sha256()
    for line in sorted(lines):
        digest.update(f"{len(line)}:".encode("utf-8"))
        digest.update(line.encode("utf-8"))
        digest.update(b"\n")
    return digest.hexdigest()


def workload_value(workload: WorkloadName, row: int, col: int, cols: int) -> str | int | None:
    if workload == "dense-values":
        return row * cols + col
    if workload == "mixed-10pct-text":
        key = row * cols + col
        return f"text-{key:08d}" if key % 10 == 0 else key
    if workload == "mixed-50pct-text":
        key = row * cols + col
        return f"text-{key:08d}" if key % 2 == 0 else key
    if workload == "mixed-closedxml-10text-5number":
        return "Hello world" if col < 10 else col - 10
    if workload == "plain-text":
        return f"text-{row * cols + col:08d}"
    if workload == "styles-heavy":
        return (row + 1) * (col + 1)
    if workload == "formula-heavy":
        base = row + 1
        if col == 0:
            return base
        if col == 1:
            return base * 2
        return base * 3 + col
    if workload == "table-heavy":
        if row == 0:
            return f"Column {col + 1}"
        if col % 3 == 0:
            return row
        if col % 3 == 1:
            return f"item-{row}-{col}"
        return row * cols + col
    if workload == "sparse-wide":
        if col == 0:
            return row
        if col == cols - 1:
            return f"edge-{row}-{cols}"
        if (row * 31 + col * 17) % 97 == 0:
            return row * cols + col
        return None
    return "Ascend" if row == 0 and col == 0 else row * cols + col


def scalar_payload(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return f"b:{str(value).lower()}"
    if isinstance(value, (int, float)):
        number = float(value)
        if number == 0:
            return "n:0"
        if number.is_integer():
            return f"n:{int(number)}"
        return f"n:{format(number, '.15g')}"
    return f"s:{value}"


def build_values(workload: WorkloadName, rows: int, cols: int) -> list[list[str | int | None]]:
    return [[workload_value(workload, row, col, cols) for col in range(cols)] for row in range(rows)]


def expected_values_hash(workload: WorkloadName, rows: int, cols: int) -> str:
    values: list[str] = []
    for row in range(rows):
        for col in range(cols):
            value = workload_value(workload, row, col, cols)
            payload = scalar_payload(value)
            if payload is not None:
                values.append(f"Data!{column_name(col + 1)}{row + 1}\t{payload}")
    return hash_lines(values)


def expected_cell_count(workload: WorkloadName, rows: int, cols: int) -> int:
    if can_validate_generated_workload(workload):
        return generated_cell_count(workload, rows, cols)
    return sum(
        1
        for row in range(rows)
        for col in range(cols)
        if workload_value(workload, row, col, cols) is not None
    )


def write_pyopenxlsx(
	workload: WorkloadName,
	rows: int,
	cols: int,
	write_strategy: str,
) -> bytes:
    import pyopenxlsx

    workbook = pyopenxlsx.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    if write_strategy == "append":
        for row in build_values(workload, rows, cols):
            sheet.append(row)
    elif write_strategy == "write-rows":
        sheet.write_rows(1, build_values(workload, rows, cols), start_col=1)
    elif write_strategy == "write-range":
        sheet.write_range(1, 1, build_values(workload, rows, cols))
    elif write_strategy == "numpy-range":
        if workload == "dense-values":
            import numpy as np

            sheet.write_range(1, 1, np.arange(rows * cols).reshape(rows, cols))
        else:
            sheet.write_rows(1, build_values(workload, rows, cols), start_col=1)
    else:
        raise ValueError(f"unsupported pyopenxlsx write strategy: {write_strategy}")
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as output:
        workbook.save(output.name)
        return Path(output.name).read_bytes()


def effective_pyopenxlsx_write_strategy(workload: WorkloadName, write_strategy: str) -> str:
    if write_strategy == "numpy-range" and workload != "dense-values":
        return "write-rows"
    return write_strategy


def write_fastxlsx(workload: WorkloadName, rows: int, cols: int) -> bytes:
    try:
        from fastxlsx import DType, WriteOnlyWorkbook
    except ModuleNotFoundError as error:
        raise RuntimeError(
            "fastxlsx is not installed; install it in a Python >=3.8,<3.13 environment"
        ) from error

    workbook = WriteOnlyWorkbook()
    sheet = workbook.create_sheet("Data")
    for row_index, row in enumerate(build_values(workload, rows, cols)):
        sheet.write_row((row_index, 0), row, dtype=DType.Any)
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as output:
        workbook.save(output.name)
        return Path(output.name).read_bytes()


def write_pyfastexcel(workload: WorkloadName, rows: int, cols: int) -> bytes:
    from pyfastexcel import Workbook

    workbook = Workbook()
    workbook.rename_sheet("Sheet1", "Data")
    sheet = workbook["Data"]
    for row_index, row in enumerate(build_values(workload, rows, cols)):
        sheet[row_index] = row
    workbook.read_lib_and_create_excel()
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as output:
        workbook.save(output.name)
        return Path(output.name).read_bytes()


def write_workbook(
    library: str,
    workload: WorkloadName,
    rows: int,
    cols: int,
    write_strategy: str,
) -> bytes:
    if library == "pyopenxlsx":
        return write_pyopenxlsx(workload, rows, cols, write_strategy)
    if library == "fastxlsx":
        return write_fastxlsx(workload, rows, cols)
    if library == "pyfastexcel":
        return write_pyfastexcel(workload, rows, cols)
    raise ValueError(f"unsupported library: {library}")


def read_fastxlsx(path: Path, rows: int, cols: int) -> tuple[str, list[str]]:
    try:
        from fastxlsx import ReadOnlyWorkbook
    except ModuleNotFoundError as error:
        raise RuntimeError(
            "fastxlsx is not installed; install it in a Python >=3.8,<3.13 environment"
        ) from error

    workbook = ReadOnlyWorkbook(str(path))
    sheet_name = workbook.sheetnames[0]
    sheet = workbook.get_by_idx(0)
    values: list[str] = []
    for row in range(rows):
        for col in range(cols):
            payload = scalar_payload(sheet.cell_value((row, col)))
            if payload is not None:
                values.append(f"{sheet_name}!{column_name(col + 1)}{row + 1}\t{payload}")
    return sheet_name, values


def read_pyopenxlsx(path: Path, rows: int, cols: int) -> tuple[str, list[str]]:
    import pyopenxlsx

    workbook = pyopenxlsx.Workbook(str(path))
    sheet = workbook.active
    sheet_name = sheet.title
    values: list[str] = []
    for row in range(rows):
        for col in range(cols):
            payload = scalar_payload(sheet.cell(row + 1, col + 1).value)
            if payload is not None:
                values.append(f"{sheet_name}!{column_name(col + 1)}{row + 1}\t{payload}")
    return sheet_name, values


def read_workbook(library: str, path: Path, rows: int, cols: int) -> tuple[str, list[str]]:
    if library == "fastxlsx":
        return read_fastxlsx(path, rows, cols)
    if library == "pyopenxlsx":
        return read_pyopenxlsx(path, rows, cols)
    raise ValueError(f"unsupported read library: {library}")


def infer_shape(path: Path) -> tuple[int, int]:
    with zipfile.ZipFile(path) as archive:
        worksheet_name = next(
            name for name in archive.namelist() if re.match(r"^xl/worksheets/sheet\d+\.xml$", name)
        )
        xml = archive.read(worksheet_name).decode("utf-8", errors="ignore")
    match = re.search(r'<dimension ref="(?:[^"]+:)?([A-Z]+)(\d+)"', xml)
    if match:
        return int(match.group(2)), column_index(match.group(1))
    rows = [int(value) for value in re.findall(r'<row[^>]*\br="(\d+)"', xml)]
    cols = [column_index(value) for value in re.findall(r'<c[^>]*\br="([A-Z]+)\d+"', xml)]
    return (max(rows) if rows else 0, max(cols) if cols else 0)


def column_index(name: str) -> int:
    value = 0
    for char in name:
        value = value * 26 + (ord(char) - 64)
    return value


def write_assertions(
	data: bytes,
	library: str,
	workload: WorkloadName,
	rows: int,
	cols: int,
	expected_hash: str | None,
	expected_ordered_hash: str | None,
	write_strategy: str,
) -> dict[str, str | int | bool | None]:
    if can_validate_generated_workload(workload):
        return {
            **generated_write_assertions(
                data,
                workload=workload,
                rows=rows,
                cols=cols,
                runner_version=runner_version(library),
                expected_ordered_hash=expected_ordered_hash,
            ),
            "requestedWriteStrategy": write_strategy,
            "effectiveWriteStrategy": (
                effective_pyopenxlsx_write_strategy(workload, write_strategy)
                if library == "pyopenxlsx"
                else write_strategy
            ),
        }
    workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=False, data_only=True)
    try:
        sheet = workbook["Data"] if "Data" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        values: list[str] = []
        for row in sheet.iter_rows():
            for cell in row:
                payload = scalar_payload(cell.value)
                if payload is not None:
                    values.append(f"Data!{column_name(cell.column)}{cell.row}\t{payload}")
        sheet_count = len(workbook.sheetnames)
    finally:
        workbook.close()
    observed_hash = hash_lines(values)
    expected_cells = expected_cell_count(workload, rows, cols)
    return {
        "runnerVersion": runner_version(library),
        "workload": workload,
        "bytes": len(data),
        "sheetCount": sheet_count,
        "cellCount": len(values),
        "expectedCellCount": expected_cells,
        "cellCountMatches": len(values) == expected_cells,
        "semanticCellValuesHash": observed_hash,
        "expectedSemanticCellValuesHash": expected_hash,
        "semanticCellValuesHashMatches": expected_hash is not None and observed_hash == expected_hash,
        "requestedWriteStrategy": write_strategy,
        "effectiveWriteStrategy": (
            effective_pyopenxlsx_write_strategy(workload, write_strategy)
            if library == "pyopenxlsx"
            else write_strategy
        ),
    }


def read_assertions(
    library: str,
    sheet_name: str,
    values: list[str],
    expected_hash: str | None = None,
) -> dict[str, str | int | bool | None]:
    semantic_hash = hash_lines(values)
    return {
        "runnerVersion": runner_version(library),
        "sheetCount": 1,
        "sheetNamesHash": hash_lines([f"0:{sheet_name}"]),
        "cellCount": len(values),
        "semanticCellValuesHash": semantic_hash,
        "expectedSemanticCellValuesHash": expected_hash,
        "semanticCellValuesHashMatches": expected_hash is None or semantic_hash == expected_hash,
        "readCommentCount": 0,
        "readHyperlinkCount": 0,
        "readDataValidationCount": 0,
        "readConditionalFormatCount": 0,
        "readDefinedNameCount": 0,
    }


def runner_version(library: str) -> str:
    try:
        return metadata.version(library)
    except metadata.PackageNotFoundError:
        module = __import__(library)
        return str(getattr(module, "__version__", "unknown"))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--library", choices=["fastxlsx", "pyopenxlsx", "pyfastexcel"], required=True)
    parser.add_argument("--operation", choices=["read", "write"], required=True)
    parser.add_argument("--file")
    parser.add_argument("--rows", type=int)
    parser.add_argument("--cols", type=int)
    parser.add_argument("--workload", choices=WORKLOAD_CHOICES, default="dense-values")
    parser.add_argument("--repeat", type=int, default=1)
    parser.add_argument("--warmup", type=int, default=0)
    parser.add_argument("--validation-mode", choices=["each", "final"], default="each")
    parser.add_argument(
        "--write-strategy",
        choices=["append", "write-rows", "write-range", "numpy-range"],
        default="append",
    )
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()

    if args.operation == "read":
        if not args.file:
            raise ValueError("--file is required for read")
        path = Path(args.file)
        rows, cols = infer_shape(path)
        for _ in range(max(0, args.warmup)):
            read_workbook(args.library, path, rows, cols)
        samples: list[dict[str, float]] = []
        assertions: dict[str, str | int | bool | None] | None = None
        for _ in range(max(1, args.repeat)):
            before = memory_baseline()
            start = time.perf_counter()
            sheet_name, values = read_workbook(args.library, path, rows, cols)
            duration_ms = (time.perf_counter() - start) * 1000
            assertions = read_assertions(args.library, sheet_name, values)
            samples.append(sample_with_memory(duration_ms, before))
        payload: dict[str, Any] = {"assertions": assertions or {}, "samples": samples}
    else:
        if args.rows is None or args.cols is None:
            raise ValueError("--rows and --cols are required for write")
        for _ in range(max(0, args.warmup)):
            write_workbook(args.library, args.workload, args.rows, args.cols, args.write_strategy)
        should_compute_expected_hash = (
            not can_validate_generated_workload(args.workload)
            or args.rows * args.cols <= 500_000
        )
        expected_hash = (
            expected_values_hash(args.workload, args.rows, args.cols)
            if should_compute_expected_hash
            else None
        )
        expected_ordered_hash = (
            expected_ordered_values_hash(args.workload, args.rows, args.cols)
            if should_compute_expected_hash
            else None
        )
        samples = []
        assertions = None
        data: bytes | None = None
        for _ in range(max(1, args.repeat)):
            before = memory_baseline()
            start = time.perf_counter()
            data = write_workbook(args.library, args.workload, args.rows, args.cols, args.write_strategy)
            duration_ms = (time.perf_counter() - start) * 1000
            if args.validation_mode == "each":
                assertions = write_assertions(
                    data,
                    args.library,
                    args.workload,
                    args.rows,
                    args.cols,
                    expected_hash,
                    expected_ordered_hash,
                    args.write_strategy,
                )
            samples.append(sample_with_memory(duration_ms, before))
        if args.validation_mode == "final":
            if data is None:
                raise RuntimeError("no workbook bytes were produced")
            assertions = write_assertions(
                data,
                args.library,
                args.workload,
                args.rows,
                args.cols,
                expected_hash,
                expected_ordered_hash,
                args.write_strategy,
            )
        if assertions is not None:
            assertions["validationMode"] = args.validation_mode
            assertions["validationSamples"] = args.repeat if args.validation_mode == "each" else 1
        payload = {"assertions": assertions or {}, "samples": samples}

    if args.json:
        print(json.dumps(payload, separators=(",", ":")))
    else:
        print(json.dumps(payload, indent=2))


if __name__ == "__main__":
    main()
