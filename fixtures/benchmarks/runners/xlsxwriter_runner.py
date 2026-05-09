#!/usr/bin/env python3
"""XlsxWriter runner for generated XLSX write benchmarks."""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import os
import tempfile
import time
import zipfile
from typing import Any
import re

import openpyxl
import xlsxwriter
from generated_xlsx_validation import (
    can_validate_generated_workload,
    expected_ordered_values_hash,
    generated_cell_count,
    generated_write_assertions,
)
from memory_metrics import memory_baseline, sample_with_memory

WorkloadName = str
WORKLOAD_CHOICES = [
    "dense-values",
    "mixed-10pct-text",
    "mixed-50pct-text",
    "mixed-closedxml-10text-5number",
    "plain-text",
    "string-heavy",
    "sparse-wide",
    "styles-heavy",
    "formula-heavy",
    "table-heavy",
    "feature-rich",
]


def column_name(index: int) -> str:
    name = ""
    current = index
    while current > 0:
        current, rem = divmod(current - 1, 26)
        name = chr(65 + rem) + name
    return name or "A"


def hash_lines(lines: list[str]) -> str:
    digest = hashlib.sha256()
    for line in sorted(lines):
        digest.update(f"{len(line)}:".encode("utf-8"))
        digest.update(line.encode("utf-8"))
        digest.update(b"\n")
    return digest.hexdigest()


def workload_value(workload: WorkloadName, row: int, col: int, cols: int) -> str | int | None:
    if workload == "dense-values":
        return row * cols + col
    if workload == "mixed-10pct-text":
        key = row * cols + col
        return f"text-{key:08d}" if key % 10 == 0 else key
    if workload == "mixed-50pct-text":
        key = row * cols + col
        return f"text-{key:08d}" if key % 2 == 0 else key
    if workload == "mixed-closedxml-10text-5number":
        return "Hello world" if col < 10 else col - 10
    if workload == "plain-text":
        return f"text-{row * cols + col:08d}"
    if workload == "styles-heavy":
        return (row + 1) * (col + 1)
    if workload == "formula-heavy":
        base = row + 1
        if col == 0:
            return base
        if col == 1:
            return base * 2
        return base * 3 + col
    if workload == "feature-rich":
        return "Ascend" if row == 0 and col == 0 else row * cols + col
    if workload == "table-heavy":
        if row == 0:
            return f"Column {col + 1}"
        if col % 3 == 0:
            return row
        if col % 3 == 1:
            return f"item-{row}-{col}"
        return row * cols + col
    if workload == "sparse-wide":
        if col == 0:
            return row
        if col == cols - 1:
            return f"edge-{row}-{cols}"
        if (row * 31 + col * 17) % 97 == 0:
            return row * cols + col
        return None
    key = row * cols + col
    branch = col % 5
    if branch == 0:
        return f"sku-{key:08d}"
    if branch == 1:
        return f"region-{(row % 17) + 1}"
    if branch == 2:
        return f"customer-{row % 997}-segment-{col % 13}"
    if branch == 3:
        return f"note row {row} col {col} token {key % 104729}"
    if key % 2 == 0:
        return f"status-open-{key % 31}"
    return f"status-closed-{key % 29}"


def build_values(workload: WorkloadName, rows: int, cols: int) -> list[list[str | int | None]]:
    return [[workload_value(workload, row, col, cols) for col in range(cols)] for row in range(rows)]


def scalar_payload(value: Any) -> str:
    if isinstance(value, bool):
        return f"b:{str(value).lower()}"
    if isinstance(value, (int, float)):
        number = float(value)
        if number == 0:
            return "n:0"
        if number.is_integer():
            return f"n:{int(number)}"
        return f"n:{format(number, '.15g')}"
    return f"s:{value}"


def workbook_formats(workbook: Any) -> list[Any]:
    return [
        None,
        workbook.add_format({"bold": True}),
        workbook.add_format({"num_format": "#,##0.00"}),
        workbook.add_format({"bg_color": "#E2F0D9"}),
        workbook.add_format({"font_color": "#1F4E79", "align": "center"}),
    ]


def write_workbook(
    workload: WorkloadName, rows: int, cols: int, constant_memory: bool = False
) -> bytes:
    style_formats: list[Any] = []

    def write_sheet(worksheet: Any) -> None:
        if workload == "sparse-wide":
            for row in range(rows):
                for col in range(cols):
                    value = workload_value(workload, row, col, cols)
                    if value is not None:
                        worksheet.write(row, col, value)
            return
        if workload == "styles-heavy":
            for row, values in enumerate(build_values(workload, rows, cols)):
                for col, value in enumerate(values):
                    worksheet.write(row, col, value, style_formats[(row + col) % len(style_formats)])
            return
        if workload == "formula-heavy":
            for row, values in enumerate(build_values(workload, rows, cols)):
                for col, value in enumerate(values):
                    if col < 2:
                        worksheet.write(row, col, value)
                    else:
                        worksheet.write_formula(row, col, f"=A{row + 1}+B{row + 1}+{col}", None, value)
            return
        for row, values in enumerate(build_values(workload, rows, cols)):
            worksheet.write_row(row, 0, values)
        if workload == "feature-rich" and rows > 0 and cols > 0:
            worksheet.write_url(
                0,
                0,
                "https://example.com/ascend",
                string="Ascend",
                tip="Open Ascend",
            )
            if rows > 1 and cols > 1:
                worksheet.write_comment(1, 1, "Review", {"author": "Ascend"})
            if rows > 1 and cols > 2:
                worksheet.data_validation(
                    1,
                    2,
                    rows - 1,
                    2,
                    {
                        "validate": "list",
                        "source": ["Q1", "Q2", "Q3"],
                        "ignore_blank": True,
                        "input_message": "Pick a quarter",
                    },
                )
            worksheet.conditional_format(
                0,
                0,
                rows - 1,
                0,
                {"type": "cell", "criteria": ">", "value": 0, "format": style_formats[1]},
            )
            workbook.define_name("FeatureRange", f"=Data!$A$1:${column_name(cols)}${rows}")
        if workload == "table-heavy" and rows > 0 and cols > 0 and not constant_memory:
            worksheet.add_table(
                0,
                0,
                rows - 1,
                cols - 1,
                {
                    "name": "DataTable",
                    "columns": [{"header": f"Column {index + 1}"} for index in range(cols)],
                },
            )

    if constant_memory:
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            workbook = xlsxwriter.Workbook(path, {"constant_memory": True})
            worksheet = workbook.add_worksheet("Data")
            style_formats = workbook_formats(workbook)
            write_sheet(worksheet)
            workbook.close()
            with open(path, "rb") as handle:
                return handle.read()
        finally:
            try:
                os.unlink(path)
            except FileNotFoundError:
                pass
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    worksheet = workbook.add_worksheet("Data")
    style_formats = workbook_formats(workbook)
    write_sheet(worksheet)
    workbook.close()
    return output.getvalue()


def expected_values_hash(workload: WorkloadName, rows: int, cols: int) -> str:
    values: list[str] = []
    for row in range(rows):
        for col in range(cols):
            ref = f"Data!{column_name(col + 1)}{row + 1}"
            value = workload_value(workload, row, col, cols)
            if value is not None:
                values.append(f"{ref}\t{scalar_payload(value)}")
    return hash_lines(values)


def expected_cell_count(workload: WorkloadName, rows: int, cols: int) -> int:
    if can_validate_generated_workload(workload):
        return generated_cell_count(workload, rows, cols)
    return sum(
        1
        for row in range(rows)
        for col in range(cols)
        if workload_value(workload, row, col, cols) is not None
    )


def write_assertions(
    data: bytes,
    workload: WorkloadName,
    rows: int,
    cols: int,
    expected_hash: str | None = None,
    expected_ordered_hash: str | None = None,
) -> dict[str, str | int | bool | None]:
    if can_validate_generated_workload(workload):
        return {
            **generated_write_assertions(
                data,
                workload=workload,
                rows=rows,
                cols=cols,
                runner_version=xlsxwriter.__version__,
                expected_ordered_hash=expected_ordered_hash,
            ),
            "formulaCount": formula_count(data),
            "tablePartCount": table_part_count(data),
            **feature_counts(data),
        }
    workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=False, data_only=True)
    try:
        sheet = workbook["Data"]
        cell_count = 0
        semantic_cell_values: list[str] = []
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell_count += 1
                ref = f"Data!{column_name(cell.column)}{cell.row}"
                semantic_cell_values.append(f"{ref}\t{scalar_payload(cell.value)}")
        observed_hash = hash_lines(semantic_cell_values)
        expected_cells = expected_cell_count(workload, rows, cols)
    finally:
        workbook.close()
    return {
        "runnerVersion": xlsxwriter.__version__,
        "workload": workload,
        "bytes": len(data),
        "formulaCount": formula_count(data),
        "tablePartCount": table_part_count(data),
        **feature_counts(data),
        "sheetCount": len(workbook.sheetnames),
        "cellCount": cell_count,
        "expectedCellCount": expected_cells,
        "cellCountMatches": cell_count == expected_cells,
        "semanticCellValuesHash": observed_hash,
        "expectedSemanticCellValuesHash": expected_hash,
        "semanticCellValuesHashMatches": expected_hash is not None and observed_hash == expected_hash,
    }


def table_part_count(data: bytes) -> int:
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            return sum(
                1
                for name in archive.namelist()
                if name.startswith("xl/tables/") and name.endswith(".xml")
            )
    except zipfile.BadZipFile:
        return 0


def formula_count(data: bytes) -> int:
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            return sum(
                len(re.findall(r"<f\b", read_zip_text(archive, name)))
                for name in archive.namelist()
                if re.match(r"^xl/worksheets/sheet\d+\.xml$", name)
            )
    except zipfile.BadZipFile:
        return 0


def feature_counts(data: bytes) -> dict[str, int]:
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            names = archive.namelist()
            workbook_xml = read_zip_text(archive, "xl/workbook.xml")
            sheet_xml = read_zip_text(archive, "xl/worksheets/sheet1.xml")
            return {
                "commentPartCount": sum(
                    1
                    for name in names
                    if re.match(r"^xl/(?:comments\d*|comments/.+)\.xml$", name)
                ),
                "vmlDrawingPartCount": sum(
                    1 for name in names if name.startswith("xl/drawings/") and name.endswith(".vml")
                ),
                "worksheetHyperlinkCount": sheet_xml.count("<hyperlink "),
                "worksheetDataValidationCount": sheet_xml.count("<dataValidation "),
                "worksheetConditionalFormattingCount": sheet_xml.count("<conditionalFormatting"),
                "definedNameCount": workbook_xml.count("<definedName "),
            }
    except zipfile.BadZipFile:
        return {
            "commentPartCount": 0,
            "vmlDrawingPartCount": 0,
            "worksheetHyperlinkCount": 0,
            "worksheetDataValidationCount": 0,
            "worksheetConditionalFormattingCount": 0,
            "definedNameCount": 0,
        }


def read_zip_text(archive: zipfile.ZipFile, name: str) -> str:
    try:
        return archive.read(name).decode("utf-8")
    except KeyError:
        return ""


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--operation", choices=["write"], required=True)
    parser.add_argument("--rows", type=int, required=True)
    parser.add_argument("--cols", type=int, required=True)
    parser.add_argument(
        "--workload", choices=WORKLOAD_CHOICES, default="dense-values"
    )
    parser.add_argument("--repeat", type=int, default=1)
    parser.add_argument("--warmup", type=int, default=0)
    parser.add_argument("--constant-memory", action="store_true")
    parser.add_argument("--validation-mode", choices=["each", "final"], default="each")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()

    for _ in range(max(0, args.warmup)):
        write_workbook(args.workload, args.rows, args.cols, args.constant_memory)
    should_compute_expected_hash = (
        not can_validate_generated_workload(args.workload)
        or args.rows * args.cols <= 500_000
    )
    expected_hash = (
        expected_values_hash(args.workload, args.rows, args.cols)
        if should_compute_expected_hash
        else None
    )
    expected_ordered_hash = (
        expected_ordered_values_hash(args.workload, args.rows, args.cols)
        if should_compute_expected_hash
        else None
    )
    samples: list[dict[str, float]] = []
    assertions: dict[str, str | int | bool | None] | None = None
    data: bytes | None = None
    for _ in range(max(1, args.repeat)):
        before = memory_baseline()
        start = time.perf_counter()
        data = write_workbook(args.workload, args.rows, args.cols, args.constant_memory)
        duration_ms = (time.perf_counter() - start) * 1000
        if args.validation_mode == "each":
            assertions = write_assertions(
                data,
                args.workload,
                args.rows,
                args.cols,
                expected_hash,
                expected_ordered_hash,
            )
            assertions["runnerMode"] = "constant-memory" if args.constant_memory else "in-memory"
        samples.append(sample_with_memory(duration_ms, before))
    if args.validation_mode == "final":
        if data is None:
            raise RuntimeError("no workbook bytes were produced")
        assertions = write_assertions(
            data,
            args.workload,
            args.rows,
            args.cols,
            expected_hash,
            expected_ordered_hash,
        )
        assertions["runnerMode"] = "constant-memory" if args.constant_memory else "in-memory"
    if assertions is not None:
        assertions["validationMode"] = args.validation_mode
        assertions["validationSamples"] = args.repeat if args.validation_mode == "each" else 1
    payload: dict[str, Any] = {"assertions": assertions or {}, "samples": samples}
    if args.json:
        print(json.dumps(payload, separators=(",", ":")))
    else:
        print(json.dumps(payload, indent=2))


if __name__ == "__main__":
    main()
